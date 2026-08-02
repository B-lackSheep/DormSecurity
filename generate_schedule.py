import asyncio
from src.database import get_async_db_session
from sqlalchemy import text
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar

RUSSIAN_MONTHS = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель",
    5: "май", 6: "июнь", 7: "июль", 8: "август",
    9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
}

ENGLISH_MONTHS = {
    1: "JANUARY", 2: "FEBRUARY", 3: "MARCH", 4: "APRIL",
    5: "MAY", 6: "JUNE", 7: "JULY", 8: "AUGUST",
    9: "SEPTEMBER", 10: "OCTOBER", 11: "NOVEMBER", 12: "DECEMBER"
}

def make_border(left=None, right=None, top=None, bottom=None):
    sides = {
        'left': Side(style=left) if left else Side(),
        'right': Side(style=right) if right else Side(),
        'top': Side(style=top) if top else Side(),
        'bottom': Side(style=bottom) if bottom else Side(),
    }
    return Border(**sides)

async def generate_duty_schedule(year, month):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(name='Calibri', size=16, bold=True)
    days_font = Font(name='Calibri', size=11, bold=False)
    room_font = Font(name='Calibri', size=11, bold=True)
    data_font = Font(name='Calibri', size=11)
    footer_font = Font(name='Calibri', size=12)
    duty_fill = openpyxl.styles.PatternFill(start_color='9BBB59', end_color='9BBB59', fill_type='solid')

    center_align = Alignment(horizontal='center', vertical='center')
    header_align = Alignment(horizontal='distributed', vertical='center')

    days_in_month = calendar.monthrange(year, month)[1]
    month_ru = RUSSIAN_MONTHS[month]
    month_en = ENGLISH_MONTHS[month]

    async with get_async_db_session() as session:
        rooms_result = await session.execute(text("""
            SELECT r.floor, r.room_number
            FROM rooms r
            ORDER BY r.floor, r.room_number
        """))
        rooms_data = rooms_result.fetchall()

        duties_result = await session.execute(text("""
            SELECT r.room_number, l.date
            FROM cleaning_log l
            JOIN rooms r ON l.room_id = r.id
            WHERE EXTRACT(YEAR FROM l.date) = :year
              AND EXTRACT(MONTH FROM l.date) = :month
        """), {"year": year, "month": month})
        duties_data = duties_result.fetchall()

    floors = {}
    for floor, room_number in rooms_data:
        if floor not in floors:
            floors[floor] = []
        floors[floor].append(room_number)

    duties_dict = {}
    for room_number, duty_date in duties_data:
        if room_number not in duties_dict:
            duties_dict[room_number] = []
        duties_dict[room_number].append(duty_date.day)

    async def fill_remaining_duties(floor_rooms, existing_duties, days_count):
        async with get_async_db_session() as session:
            last_duties_result = await session.execute(text("""
                SELECT r.room_number, COALESCE(MAX(l.date), '1900-01-01') as last_date
                FROM rooms r
                LEFT JOIN cleaning_log l ON l.room_id = r.id
                WHERE r.room_number = ANY(:room_numbers)
                GROUP BY r.room_number
                ORDER BY last_date ASC, r.room_number
            """), {"room_numbers": floor_rooms})
            last_duties_data = last_duties_result.fetchall()

        room_priority = [room_num for room_num, _ in last_duties_data]
        filled_duties = existing_duties.copy()

        occupied_days = set()
        for room_duties in filled_duties.values():
            occupied_days.update(room_duties)

        empty_days = [day for day in range(1, days_count + 1) if day not in occupied_days]

        room_index = 0
        for day in empty_days:
            room_number = room_priority[room_index % len(room_priority)]
            if room_number not in filled_duties:
                filled_duties[room_number] = []
            filled_duties[room_number].append(day)
            room_index += 1

        return filled_duties

    for floor in sorted(floors.keys()):
        ws = wb.create_sheet(title=f"Этаж {floor}")
        rooms_on_floor = sorted(floors[floor])

        floor_duties = {room: duties_dict.get(room, []) for room in rooms_on_floor}
        complete_duties = await fill_remaining_duties(rooms_on_floor, floor_duties, days_in_month)

        # Column layout: A = room numbers, B...(B+days_in_month-1) = day columns
        last_day_col = get_column_letter(1 + days_in_month)  # e.g. AF for 31 days

        # --- Row 1: Header ---
        ws.merge_cells(f'A1:{last_day_col}1')
        ws['A1'] = (
            f"ГРАФИК ДЕЖУРСТВА НА {floor} этаже ({month_ru})| "
            f"DUTY SCHEDULE OF ROOMS ON THE FLOOR ({month_en})"
        )
        ws['A1'].font = header_font
        ws['A1'].alignment = header_align
        ws['A1'].border = make_border(left='medium', right='medium', top='medium', bottom='medium')
        ws.row_dimensions[1].height = 45

        # --- Row 2: Day numbers ---
        # A2 is empty (corner cell above room numbers column)
        ws['A2'].border = make_border(left='medium', right='medium', top='medium', bottom='medium')

        for day in range(1, days_in_month + 1):
            col = get_column_letter(1 + day)
            cell = ws[f'{col}2']
            cell.value = day
            cell.font = days_font
            cell.alignment = center_align
            if day == 1:
                cell.border = make_border(left='medium', right='thin', top='medium', bottom='medium')
            elif day == days_in_month:
                cell.border = make_border(left='medium', right='medium', top='medium', bottom='medium')
            else:
                cell.border = make_border(left='thin', right='thin', top='medium', bottom='medium')

        ws.row_dimensions[2].height = 15.25

        # --- Rows 3..N: Room rows ---
        for row_idx, room_number in enumerate(rooms_on_floor, start=3):
            # Column A: room number
            cell_a = ws[f'A{row_idx}']
            cell_a.value = room_number
            cell_a.font = room_font
            cell_a.alignment = center_align
            cell_a.border = make_border(left='medium', right='medium', bottom='thin')

            # Day columns
            for day in range(1, days_in_month + 1):
                col = get_column_letter(1 + day)
                cell = ws[f'{col}{row_idx}']
                # Mark duty day with "+" or leave empty
                if room_number in complete_duties and day in complete_duties[room_number]:
                    cell.fill = duty_fill
                cell.font = data_font
                cell.alignment = center_align
                if day == 1:
                    cell.border = make_border(left='medium', right='thin', bottom='thin')
                elif day == days_in_month:
                    cell.border = make_border(left='medium', right='medium', bottom='thin')
                else:
                    cell.border = make_border(left='thin', right='thin', bottom='thin')

        # --- Footer row ---
        footer_row = len(rooms_on_floor) + 3
        ws.merge_cells(f'A{footer_row}:{last_day_col}{footer_row}')
        ws[f'A{footer_row}'].value = f"Ответственный за составление графика – Карачун Е. В."
        ws[f'A{footer_row}'].font = footer_font
        ws[f'A{footer_row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[footer_row].height = 16.25

        # --- Column widths ---
        ws.column_dimensions['A'].width = 7.45
        for day in range(1, days_in_month + 1):
            col = get_column_letter(1 + day)
            ws.column_dimensions[col].width = 3.64

    return wb


async def main():
    print("📅 ГЕНЕРАТОР ГРАФИКА ДЕЖУРСТВ")
    print("=" * 40)

    try:
        year = int(input("Введите год (например, 2026): "))
        month = int(input("Введите месяц (1-12): "))
        if not (1 <= month <= 12):
            print("❌ Месяц должен быть от 1 до 12")
            return
    except ValueError:
        print("❌ Неверный формат. Используйте числа.")
        return

    print(f"\n📊 Генерирую график дежурств за {RUSSIAN_MONTHS[month]} {year}...")

    wb = await generate_duty_schedule(year, month)

    filename = f"График_дежурств_{calendar.month_name[month]}_{year}.xlsx"
    wb.save(filename)

    print(f"✅ Файл сохранён: {filename}")

if __name__ == "__main__":
    asyncio.run(main())
